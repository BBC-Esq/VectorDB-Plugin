import os
import csv
import logging
import warnings
import datetime
import hashlib
import re
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Tuple, Optional
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed

import fitz
from bs4 import BeautifulSoup

from core.utilities import normalize_text
from core.constants import SUPPORTED_EXTENSIONS, PIPELINE_PRESETS

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)

THREADS_PER_PROCESS = 4


def _get_ingest_params():
    try:
        from core.config import get_config
        preset_name = get_config().database.pipeline_preset
    except Exception:
        preset_name = "normal"
    preset = PIPELINE_PRESETS.get(preset_name, PIPELINE_PRESETS["normal"])
    return preset["ingest_threads"], preset["ingest_processes"]

logger = logging.getLogger(__name__)


@dataclass
class Document:
    page_content: str = ""
    metadata: dict = field(default_factory=dict)


def compute_content_hash(content: str) -> str:
    return hashlib.sha256(content.encode('utf-8')).hexdigest()


def compute_file_hash(file_path):
    hash_sha256 = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_sha256.update(chunk)
    return hash_sha256.hexdigest()


def extract_document_metadata(file_path, content_hash=None):
    file_path = os.path.realpath(file_path)
    file_name = os.path.basename(file_path)
    file_type = os.path.splitext(file_path)[1]
    creation_date = datetime.datetime.fromtimestamp(os.path.getctime(file_path)).isoformat()
    modification_date = datetime.datetime.fromtimestamp(os.path.getmtime(file_path)).isoformat()
    file_hash = content_hash if content_hash else compute_file_hash(file_path)

    return {
        "file_path": file_path,
        "file_type": file_type,
        "file_name": file_name,
        "creation_date": creation_date,
        "modification_date": modification_date,
        "hash": file_hash,
        "document_type": "document",
    }


def _load_pdf(file_path: Path) -> Optional[str]:
    full_content = []
    with fitz.open(str(file_path)) as doc:
        for page in doc:
            text = page.get_text()
            if text.strip():
                full_content.append(f"[[page{page.number + 1}]]{text}")
    return "".join(full_content) if full_content else None


def _load_docx(file_path: Path) -> Optional[str]:
    import docx2txt
    text = docx2txt.process(str(file_path))
    return text if text and text.strip() else None


def _load_txt(file_path: Path) -> Optional[str]:
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                text = f.read()
            return text if text and text.strip() else None
        except UnicodeDecodeError:
            continue
    return None


def _load_csv(file_path: Path) -> Optional[str]:
    rows = []
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc, newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    rows.append(" ".join(row))
            break
        except UnicodeDecodeError:
            continue
    return "\n".join(rows) if rows else None


def _load_html(file_path: Path) -> Optional[str]:
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                soup = BeautifulSoup(f, "lxml")
            text = soup.get_text(separator=" ")
            return text if text and text.strip() else None
        except UnicodeDecodeError:
            continue
    return None


def _load_eml(file_path: Path) -> Optional[str]:
    import email
    from email import policy

    with open(file_path, "rb") as f:
        msg = email.message_from_binary_file(f, policy=policy.default)

    parts = []
    subject = msg.get("Subject", "")
    if subject:
        parts.append(f"Subject: {subject}")

    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            if content_type == "text/plain":
                payload = part.get_content()
                if isinstance(payload, str) and payload.strip():
                    parts.append(payload)
            elif content_type == "text/html":
                payload = part.get_content()
                if isinstance(payload, str):
                    soup = BeautifulSoup(payload, "lxml")
                    text = soup.get_text(separator=" ")
                    if text.strip():
                        parts.append(text)
    else:
        payload = msg.get_content()
        if isinstance(payload, str) and payload.strip():
            parts.append(payload)

    return "\n".join(parts) if parts else None


def _load_msg(file_path: Path) -> Optional[str]:
    import extract_msg

    msg = extract_msg.Message(str(file_path))
    parts = []
    if msg.subject:
        parts.append(f"Subject: {msg.subject}")
    if msg.body:
        parts.append(msg.body)
    msg.close()
    return "\n".join(parts) if parts else None


def _load_xls(file_path: Path) -> Optional[str]:
    import xlrd

    workbook = xlrd.open_workbook(str(file_path))
    parts = []
    for sheet in workbook.sheets():
        for row_idx in range(sheet.nrows):
            row_values = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                if cell.value is not None and str(cell.value).strip():
                    row_values.append(str(cell.value))
            if row_values:
                parts.append(" ".join(row_values))
    return "\n".join(parts) if parts else None


def _load_xlsx(file_path: Path) -> Optional[str]:
    from openpyxl import load_workbook

    wb = load_workbook(str(file_path), data_only=True, read_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            row_values = []
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    row_values.append(str(cell.value))
            if row_values:
                parts.append(" ".join(row_values))
    wb.close()
    return "\n".join(parts) if parts else None


def _load_rtf(file_path: Path) -> Optional[str]:
    from striprtf.striprtf import rtf_to_text

    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                rtf_content = f.read()
            text = rtf_to_text(rtf_content)
            return text if text and text.strip() else None
        except UnicodeDecodeError:
            continue
    return None


def _load_md(file_path: Path) -> Optional[str]:
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                text = f.read()
            return text if text and text.strip() else None
        except UnicodeDecodeError:
            continue
    return None


LOADER_MAP = {
    ".pdf": _load_pdf,
    ".docx": _load_docx,
    ".txt": _load_txt,
    ".csv": _load_csv,
    ".html": _load_html,
    ".htm": _load_html,
    ".eml": _load_eml,
    ".msg": _load_msg,
    ".xls": _load_xls,
    ".xlsx": _load_xlsx,
    ".xlsm": _load_xlsx,
    ".rtf": _load_rtf,
    ".md": _load_md,
}


def load_single_document(file_path: Path) -> Optional[Document]:
    file_extension = file_path.suffix.lower()
    loader_fn = LOADER_MAP.get(file_extension)

    if not loader_fn:
        print(f"\033[91mFailed---> {file_path.name} (extension: {file_extension})\033[0m")
        logger.error(f"Unsupported file type: {file_path.name} (extension: {file_extension})")
        return None

    try:
        content = loader_fn(file_path)

        if not content:
            print(f"\033[91mFailed---> {file_path.name} (No content extracted)\033[0m")
            logger.error(f"No content extracted: {file_path.name}")
            return None

        content_hash = compute_content_hash(content)
        metadata = extract_document_metadata(file_path, content_hash)
        print(f"Loaded---> {file_path.name}")
        return Document(page_content=content, metadata=metadata)

    except (OSError, UnicodeDecodeError) as e:
        print(f"\033[91mFailed---> {file_path.name} (Access/encoding error)\033[0m")
        logger.error(f"File access/encoding error - File: {file_path.name} - Error: {str(e)}")
        return None
    except Exception as e:
        print(f"\033[91mFailed---> {file_path.name} (Unexpected error)\033[0m")
        logger.error(f"Unexpected error processing file: {file_path.name} - Error: {type(e).__name__}: {str(e)}")
        logging.exception("Full traceback:")
        return None


def _extraction_worker_batch(file_paths):
    results = []

    def _process_one(file_path):
        return load_single_document(file_path)

    n_threads = min(THREADS_PER_PROCESS, len(file_paths))
    with ThreadPoolExecutor(n_threads) as pool:
        futures = {pool.submit(_process_one, p): p for p in file_paths}
        for future in as_completed(futures):
            try:
                doc = future.result()
                if doc is not None:
                    results.append((doc.page_content, doc.metadata))
            except Exception as e:
                path = futures[future]
                logger.error(f"Error processing document {path}: {e}")

    return results


def load_documents(source_dir: Path) -> list:
    valid_extensions = set(SUPPORTED_EXTENSIONS)
    doc_paths = [f for f in source_dir.iterdir() if f.suffix.lower() in valid_extensions]

    docs = []

    if not doc_paths:
        return docs

    ingest_threads, ingest_processes = _get_ingest_params()

    if len(doc_paths) <= ingest_processes:
        n_workers = min(ingest_threads, max(len(doc_paths), 1))

        executor = None
        try:
            executor = ThreadPoolExecutor(n_workers)
            futures = [executor.submit(load_single_document, path) for path in doc_paths]
            for future in as_completed(futures):
                try:
                    result = future.result()
                    if result is not None:
                        docs.append(result)
                except Exception as e:
                    logger.error(f"Error processing document: {e}")
        except Exception as e:
            logger.error(f"Error in document loading executor: {e}")
            raise
        finally:
            if executor:
                executor.shutdown(wait=True, cancel_futures=True)
    else:
        n_procs = min(ingest_processes, len(doc_paths))
        logger.info(f"Loading {len(doc_paths)} documents with {n_procs} processes \u00b7 {THREADS_PER_PROCESS} threads each")

        chunks = [[] for _ in range(n_procs)]
        for i, chunk in enumerate(doc_paths):
            chunks[i % n_procs].append(chunk)

        try:
            with ProcessPoolExecutor(n_procs) as executor:
                futures = [executor.submit(_extraction_worker_batch, chunk) for chunk in chunks]
                for future in as_completed(futures):
                    try:
                        batch_results = future.result()
                        for content, metadata in batch_results:
                            docs.append(Document(page_content=content, metadata=metadata))
                    except Exception as e:
                        logger.error(f"Error in extraction worker: {e}")
        except Exception as e:
            logger.error(f"Error in multi-process document loading: {e}")
            raise

    return docs


class FixedSizeTextSplitter:
    def __init__(self, chunk_size: int, chunk_overlap: int = 0):
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap

    def split_documents(self, docs: List[Document]) -> List[Document]:
        chunks: List[Document] = []
        step = self.chunk_size - self.chunk_overlap
        if step <= 0:
            step = 1

        for doc in docs:
            text = doc.page_content

            if text is None:
                logger.warning("Skipping document with None page_content")
                continue

            if isinstance(text, (list, tuple)):
                text = " ".join(str(item) for item in text if item)
                logger.warning("Flattened list/tuple page_content to string")

            if not isinstance(text, str):
                text = str(text)

            text = text.strip()

            if not text:
                logger.warning("Skipping document with empty page_content")
                continue

            for start in range(0, len(text), step):
                piece = text[start:start + self.chunk_size].strip()

                if not piece:
                    continue

                metadata = doc.metadata if doc.metadata else {}
                chunks.append(Document(page_content=piece, metadata=dict(metadata)))

        return chunks


def add_pymupdf_page_metadata(doc: Document, chunk_size: int = 1200, chunk_overlap: int = 600) -> List[Document]:
    def split_text(text: str, chunk_size: int, chunk_overlap: int) -> List[Tuple[str, int]]:
        if text is None:
            return []

        if isinstance(text, (list, tuple)):
            text = " ".join(str(item) for item in text if item)

        if not isinstance(text, str):
            text = str(text)

        page_markers = []
        offset = 0
        for m in re.finditer(r'\[\[page(\d+)\]\]', text):
            marker_len = len(m.group(0))
            page_markers.append((m.start() - offset, int(m.group(1))))
            offset += marker_len

        clean_text = re.sub(r'\[\[page\d+\]\]', '', text)

        chunks = []
        start = 0
        while start < len(clean_text):
            end = start + chunk_size
            if end > len(clean_text):
                end = len(clean_text)
            chunk = clean_text[start:end].strip()

            page_num = None
            for marker_pos, page in reversed(page_markers):
                if marker_pos <= start:
                    page_num = page
                    break

            if chunk and page_num is not None:
                chunks.append((chunk, page_num))
            elif chunk and page_num is None:
                chunks.append((chunk, 1))

            start += chunk_size - chunk_overlap

        return chunks

    text = doc.page_content

    if text is None:
        logger.warning("Skipping PDF document with None page_content")
        return []

    chunks = split_text(text, chunk_size, chunk_overlap)

    if not chunks:
        logger.warning("No chunks created from PDF document")
        return []

    new_docs = []
    for chunk, page_num in chunks:
        if not chunk or not chunk.strip():
            continue

        new_metadata = doc.metadata.copy() if doc.metadata else {}
        new_metadata['page_number'] = page_num

        new_doc = Document(page_content=chunk, metadata=new_metadata)
        new_docs.append(new_doc)

    return new_docs


def split_documents(documents=None, text_documents_pdf=None, chunk_size=None, chunk_overlap=None):
    try:
        print("\nSplitting documents into chunks.")

        if chunk_size is None or chunk_overlap is None:
            from core.config import get_config
            config = get_config()
            chunk_size = chunk_size if chunk_size is not None else config.database.chunk_size
            chunk_overlap = chunk_overlap if chunk_overlap is not None else config.database.chunk_overlap

        text_splitter = FixedSizeTextSplitter(chunk_size=chunk_size, chunk_overlap=chunk_overlap)

        texts = []

        if documents:
            texts = text_splitter.split_documents(documents)

        if text_documents_pdf:
            processed_pdf_docs = []
            for doc in text_documents_pdf:
                chunked_docs = add_pymupdf_page_metadata(
                    doc,
                    chunk_size=chunk_size,
                    chunk_overlap=chunk_overlap,
                )
                processed_pdf_docs.extend(chunked_docs)
            texts.extend(processed_pdf_docs)

        normalized = []
        for doc in texts:
            cleaned = normalize_text(doc.page_content, preserve_whitespace=True)
            if cleaned is None:
                logger.warning(f"Dropping chunk with empty content after normalization "
                               f"(source: {doc.metadata.get('file_name', 'unknown')})")
                continue
            doc.page_content = cleaned
            normalized.append(doc)

        texts = normalized
        print(f"Total chunks after splitting and normalization: {len(texts)}")

        return texts

    except Exception as e:
        logging.exception("Error during document splitting")
        logger.error(f"Error type: {type(e)}")
        raise
